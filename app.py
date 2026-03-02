import streamlit as st
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import io
import re
from PIL import Image
from datetime import datetime

st.set_page_config(page_title="Rechnung Excel", page_icon="", layout="centered")

SHEET_CONFIG = {
    "Baustellenzubehoer": {
        "data_start": 6, "sum_row": 30, "rn_col": 2, "datum_col": 3,
        "betrag_col": 4, "brutto_col": 5,
        "keywords": ["werkzeug", "geruest", "container", "baustelleneinrichtung",
                      "schutz", "kleinmaterial", "sicherheit", "absperrung"]
    },
    "Demontage": {
        "data_start": 7, "sum_row": 28, "rn_col": 2, "datum_col": 3,
        "betrag_col": 4, "brutto_col": 5,
        "keywords": ["abbruch", "rueckbau", "entsorgung", "demontage", "abriss"]
    },
    "Trockenbau": {
        "data_start": 6, "sum_row": 21, "rn_col": 2, "datum_col": 3,
        "betrag_col": 4, "brutto_col": 5,
        "keywords": ["trockenbau", "rigips", "staenderwerk", "gipskarton",
                      "spachtel", "fassade", "daemmung", "innenausbau", "putz"]
    },
    "Nordrheinbau": {
        "data_start": 2, "sum_row": 10, "rn_col": 2, "datum_col": 3,
        "betrag_col": 4, "brutto_col": 5,
        "keywords": ["nordrheinbau", "nordrhein"]
    },
    "Lieferanten": {
        "data_start": 6, "sum_row": 16, "rn_col": 2, "datum_col": 3,
        "betrag_col": 4, "brutto_col": 5,
        "keywords": ["material", "baustoffe", "farbe", "fliesen", "sanitaer",
                      "elektro", "lieferung", "lieferant", "holz", "schrauben"]
    },
}


def extract_invoice_data(image_bytes, api_key):
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-1.5-flash-latest")
    img = Image.open(io.BytesIO(image_bytes))
    prompt = """Analysiere diese Rechnung und extrahiere folgende Daten als JSON.
Antworte NUR mit dem JSON-Objekt, ohne Markdown-Codebloecke oder sonstigen Text.

{
  "rechnungsnummer": "...",
  "datum": "TT.MM.JJJJ",
  "netto": 0.00,
  "brutto": 0.00,
  "firma": "...",
  "beschreibung": "kurze Beschreibung der Leistung/Lieferung"
}

Wichtig:
- Datum im Format TT.MM.JJJJ
- Betraege als Zahlen ohne Waehrungszeichen
- Netto = ohne MwSt, Brutto = mit MwSt
- Wenn nur ein Betrag vorhanden ist, setze ihn als Brutto und berechne Netto (/ 1.19)
"""
    response = model.generate_content([prompt, img])
    text = response.text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


def suggest_sheet(beschreibung, firma):
    text = (beschreibung + " " + firma).lower()
    best_sheet = "Lieferanten"
    best_score = 0
    for sheet, cfg in SHEET_CONFIG.items():
        score = sum(1 for kw in cfg["keywords"] if kw in text)
        if score > best_score:
            best_score = score
            best_sheet = sheet
    return best_sheet


def find_next_empty_row(ws, config):
    for row in range(config["data_start"], config["sum_row"]):
        rn_val = ws.cell(row=row, column=config["rn_col"]).value
        betrag_val = ws.cell(row=row, column=config["betrag_col"]).value
        if rn_val is None and betrag_val is None:
            return row
    return None


def enter_invoice(excel_bytes, sheet_name, rn, datum, netto, brutto):
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb[sheet_name]
    config = SHEET_CONFIG[sheet_name]
    target_row = find_next_empty_row(ws, config)
    if target_row is None:
        raise ValueError(f"Kein Platz mehr im Sheet '{sheet_name}'.")
    ws.cell(row=target_row, column=config["rn_col"], value=rn)
    ws.cell(row=target_row, column=config["datum_col"], value=datum)
    netto_cell = ws.cell(row=target_row, column=config["betrag_col"], value=netto)
    netto_cell.number_format = '#,##0.00'
    brutto_cell = ws.cell(row=target_row, column=config["brutto_col"], value=brutto)
    brutto_cell.number_format = '#,##0.00'
    header_row = 2 if sheet_name != "Nordrheinbau" else 1
    hdr = ws.cell(row=header_row, column=config["brutto_col"])
    if hdr.value is None:
        hdr.value = "Brutto"
        hdr.font = Font(bold=True)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


st.title("Rechnung - Excel")
st.caption("Lade eine Rechnung hoch und sie wird automatisch in die Kosten-Excel eingetragen.")

with st.sidebar:
    st.header("Einstellungen")
    api_key = st.text_input("Gemini API Key", type="password",
                            help="Hol dir einen kostenlosen Key auf ai.google.dev")
    st.divider()
    st.subheader("Excel-Datei")
    excel_file = st.file_uploader("Kosten-Excel hochladen", type=["xlsx"],
                                   help="Die Datei Kosten intern AOK.xlsx")
    if excel_file:
        st.success(f"Geladen: {excel_file.name}")

if not api_key:
    st.info("Bitte gib deinen **Gemini API Key** in der Seitenleiste ein, um zu starten.")
    st.markdown("Kostenlosen Key holen: [ai.google.dev](https://ai.google.dev)")
    st.stop()

if not excel_file:
    st.info("Bitte lade die **Kosten-Excel** in der Seitenleiste hoch.")
    st.stop()

excel_bytes = excel_file.read()

st.subheader("Rechnung hochladen")
invoice_file = st.file_uploader("Rechnungsbild oder PDF", type=["jpg", "jpeg", "png", "pdf", "webp"])

if invoice_file:
    if invoice_file.type.startswith("image"):
        st.image(invoice_file, caption="Hochgeladene Rechnung", use_container_width=True)
    invoice_bytes = invoice_file.read()
    if st.button("Rechnung analysieren", type="primary", use_container_width=True):
        with st.spinner("Rechnung wird analysiert..."):
            try:
                data = extract_invoice_data(invoice_bytes, api_key)
                st.session_state["invoice_data"] = data
                st.session_state["suggested_sheet"] = suggest_sheet(
                    data.get("beschreibung", ""), data.get("firma", ""))
            except Exception as e:
                st.error(f"Fehler bei der Analyse: {e}")

if "invoice_data" in st.session_state:
    data = st.session_state["invoice_data"]
    suggested = st.session_state.get("suggested_sheet", "Lieferanten")
    st.divider()
    st.subheader("Erkannte Daten")
    col1, col2 = st.columns(2)
    with col1:
        rn = st.text_input("Rechnungsnummer", value=data.get("rechnungsnummer", ""))
        datum = st.text_input("Datum (TT.MM.JJJJ)", value=data.get("datum", ""))
        firma = st.text_input("Firma", value=data.get("firma", ""), disabled=True)
    with col2:
        netto = st.number_input("Netto", value=float(data.get("netto", 0)),
                                 format="%.2f", step=0.01)
        brutto = st.number_input("Brutto", value=float(data.get("brutto", 0)),
                                  format="%.2f", step=0.01)
        beschreibung = st.text_input("Beschreibung", value=data.get("beschreibung", ""),
                                      disabled=True)
    sheet_names = list(SHEET_CONFIG.keys())
    default_idx = sheet_names.index(suggested) if suggested in sheet_names else 4
    sheet = st.selectbox("Gewerk / Sheet", sheet_names, index=default_idx,
                         help="Automatisch vorgeschlagen, kann geaendert werden")
    st.divider()
    if st.button("In Excel eintragen", type="primary", use_container_width=True):
        try:
            updated = enter_invoice(excel_bytes, sheet, rn, datum, netto, brutto)
            st.success("Rechnung erfolgreich eingetragen!")
            st.markdown(f"**{rn}** vom {datum}")
            st.markdown(f"Netto: **{netto:,.2f}** / Brutto: **{brutto:,.2f}**")
            st.markdown(f"Eingetragen in: **{sheet}**")
            st.download_button(
                "Aktualisierte Excel herunterladen",
                data=updated,
                file_name=f"Kosten_intern_AOK_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Fehler beim Eintragen: {e}")

st.divider()
st.caption("Rechnungs-Automation - Powered by Gemini AI + Streamlit")
